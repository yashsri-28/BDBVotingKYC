"""
Imports the Category Trade Member and Exclusive Member electoral rolls
into ElectoralRoll (MoM: Voting Module Review, 2026-07-24).

Ballot logic confirmed 2026-07-27:
  - Category roll "Category" column: I -> 1 ballot, II -> 2, III -> 3
    (asterisk suffixes on the Exclusive roll's Membership Number are
    confirmed meaningless -- ignored on import)
  - Exclusive roll: flat 1 ballot per member (no tier column exists)

Being on either roll means a member is IN SCOPE for this election --
it does NOT mean they are automatically eligible. Eligibility is still
decided by the Section 5 business rules at the counter, same as before.

Usage:
    python manage.py import_electoral_rolls
    python manage.py import_electoral_rolls --category-file path/to/file.xlsx --exclusive-file path/to/file.xlsx
"""
import re
import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.conf import settings

from apps.ballots.models import ElectoralRoll, RollType
from apps.kyc_portal.models import MembersMaster

DEFAULT_CATEGORY_FILE = settings.BASE_DIR / "data" / "Electoral_Roll__Category_Trade_Member.xlsx"
DEFAULT_EXCLUSIVE_FILE = settings.BASE_DIR / "data" / "Electoral_Rol_Exclusive.xlsx"

ASTERISK_RE = re.compile(r"\*+$")


class Command(BaseCommand):
    help = "Import the Category Trade Member and Exclusive Member electoral rolls."

    def add_arguments(self, parser):
        parser.add_argument("--category-file", default=str(DEFAULT_CATEGORY_FILE))
        parser.add_argument("--exclusive-file", default=str(DEFAULT_EXCLUSIVE_FILE))
        parser.add_argument(
            "--clear", action="store_true",
            help="Delete all existing ElectoralRoll rows before importing. "
                 "Recommended for re-imports, since membership numbers are "
                 "cleaned (asterisks stripped) and won't match old dirty rows.",
        )

    def handle(self, *args, **options):
        if options["clear"]:
            deleted, _ = ElectoralRoll.objects.all().delete()
            self.stdout.write(self.style.WARNING(f"Cleared {deleted} existing electoral roll rows."))
        self.import_category(options["category_file"])
        self.import_exclusive(options["exclusive_file"])
        self.stdout.write(self.style.SUCCESS("Electoral roll import complete."))

    def _match_customer_code(self, membership_no):
        """Best-effort match against the real KYC data so downstream
        eligibility checks can join straight through. Import still
        succeeds even with no match -- customer_code stays null and can
        be reconciled later."""
        member = MembersMaster.objects.filter(membership_no=membership_no).first()
        return member.customer_code if member else None

    def import_category(self, path):
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except FileNotFoundError:
            raise CommandError(f"Category roll not found at {path}")
        ws = wb["Sheet1"]

        created, updated, unmatched = 0, 0, 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            _, tm_no, name, category, _shares, auth_person, *_ = row
            if not tm_no:
                continue
            # Asterisk suffixes are confirmed meaningless -- strip them (same as the Exclusive roll).
            tm_no = ASTERISK_RE.sub("", str(tm_no).strip())
            tier = str(category).strip() if category else None
            if tier not in ("I", "II", "III"):
                self.stdout.write(self.style.WARNING(f"Skipping {tm_no}: unrecognized category '{category}'"))
                continue

            customer_code = self._match_customer_code(tm_no)
            if not customer_code:
                unmatched += 1

            _, was_created = ElectoralRoll.objects.update_or_create(
                roll_type=RollType.CATEGORY, membership_no=tm_no,
                defaults=dict(
                    customer_code=customer_code,
                    entity_name=str(name).strip() if name else "",
                    representative_name=str(auth_person).strip() if auth_person else "",
                    category_tier=tier,
                ),
            )
            created += was_created
            updated += not was_created

        self.stdout.write(self.style.SUCCESS(
            f"Category roll: {created} created, {updated} updated, {unmatched} not matched to KYC data."
        ))

    def import_exclusive(self, path):
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except FileNotFoundError:
            raise CommandError(f"Exclusive roll not found at {path}")
        ws = wb["Sheet1"]

        created, updated, unmatched = 0, 0, 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            _, membership_no, name, rep_name, email1, *_ = row
            if not membership_no:
                continue
            # Asterisk suffixes are confirmed meaningless -- strip them.
            clean_no = ASTERISK_RE.sub("", str(membership_no).strip())

            customer_code = self._match_customer_code(clean_no)
            if not customer_code:
                unmatched += 1

            _, was_created = ElectoralRoll.objects.update_or_create(
                roll_type=RollType.EXCLUSIVE, membership_no=clean_no,
                defaults=dict(
                    customer_code=customer_code,
                    entity_name=str(name).strip() if name else "",
                    representative_name=str(rep_name).strip() if rep_name else "",
                    representative_email=str(email1).strip() if email1 else "",
                    category_tier=None,
                ),
            )
            created += was_created
            updated += not was_created

        self.stdout.write(self.style.SUCCESS(
            f"Exclusive roll: {created} created, {updated} updated, {unmatched} not matched to KYC data."
        ))
